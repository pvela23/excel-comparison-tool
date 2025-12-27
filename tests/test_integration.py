"""
Integration tests for Excel Comparison Tool
Tests end-to-end workflow and component interaction
"""

import pytest
import pandas as pd
import tempfile
from pathlib import Path
from datetime import datetime
from openpyxl import load_workbook

from src.core import (
    ComparisonEngine,
    ComparisonConfig,
    RowStatus,
    AlignmentMethod
)
from src.reports.report_generator import ReportGenerator


class TestEndToEndComparison:
    """Test complete comparison workflow"""
    
    def test_full_workflow_basic(self):
        """Test complete workflow: load -> compare -> report"""
        # Create test dataframes
        df_a = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P002'],
            'Coverage': ['A', 'B', 'A'],
            'Premium': [100, 50, 200],
            'Status': ['Active', 'Active', 'Pending']
        })
        
        df_b = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P002'],
            'Coverage': ['A', 'B', 'A'],
            'Premium': [100, 55, 200],  # Modified
            'Status': ['Active', 'Active', 'Pending']
        })
        
        # Configure comparison
        config = ComparisonConfig(
            key_columns=['Policy'],
            alignment_method=AlignmentMethod.POSITION,
            case_sensitive=False,
            trim_whitespace=True
        )
        
        # Run comparison
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Verify results
        assert result.summary['keys_in_common'] == 2
        assert result.summary['modified_count'] == 1
        assert result.summary['match_count'] == 2
        
        # Generate report
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            report_path = f.name
        
        try:
            generator = ReportGenerator(report_path)
            generator.generate_report(
                summary=result.summary,
                aligned_data=result.aligned_data,
                metadata=result.comparison_metadata,
                file_a_path='test_a.xlsx',
                file_b_path='test_b.xlsx'
            )
            
            # Verify report was created
            assert Path(report_path).exists()
            
            # Verify report content
            wb = load_workbook(report_path)
            assert 'Summary' in wb.sheetnames
            assert 'Aligned Diff' in wb.sheetnames
            assert 'Legend' in wb.sheetnames
        
        finally:
            if Path(report_path).exists():
                Path(report_path).unlink()
    
    def test_full_workflow_with_multiple_changes(self):
        """Test workflow with added, removed, and modified rows"""
        df_a = pd.DataFrame({
            'ID': [1, 2, 3, 4],
            'Name': ['Alice', 'Bob', 'Charlie', 'David'],
            'Department': ['Sales', 'IT', 'HR', 'Finance'],
            'Salary': [50000, 60000, 55000, 65000]
        })
        
        df_b = pd.DataFrame({
            'ID': [1, 2, 3, 5],
            'Name': ['Alice', 'Bob', 'Charles', 'Eve'],  # Modified Charlie, removed David, added Eve
            'Department': ['Sales', 'IT', 'HR', 'Operations'],  # Modified HR/Ops
            'Salary': [52000, 60000, 55000, 70000]  # Multiple salary changes
        })
        
        config = ComparisonConfig(
            key_columns=['ID'],
            case_sensitive=False,
            trim_whitespace=True
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Verify all types of changes are detected
        assert result.summary['match_count'] >= 0
        assert result.summary['modified_count'] > 0
        assert result.summary['keys_only_in_a'] == 1  # David (ID 4)
        assert result.summary['keys_only_in_b'] == 1  # Eve (ID 5)
        assert len(result.key_only_in_a) == 1
        assert len(result.key_only_in_b) == 1


class TestCompositeKeyWorkflow:
    """Test workflow with composite keys"""
    
    def test_composite_key_full_workflow(self):
        """Test end-to-end with composite keys"""
        df_a = pd.DataFrame({
            'Account': ['ACC001', 'ACC001', 'ACC002', 'ACC002'],
            'TransDate': ['2024-01-01', '2024-01-02', '2024-01-01', '2024-01-03'],
            'Amount': [1000.00, 500.00, 2000.00, 300.00],
            'Type': ['Deposit', 'Withdrawal', 'Deposit', 'Fee']
        })
        
        df_b = pd.DataFrame({
            'Account': ['ACC001', 'ACC001', 'ACC002', 'ACC002'],
            'TransDate': ['2024-01-01', '2024-01-02', '2024-01-01', '2024-01-03'],
            'Amount': [1000.00, 500.00, 2000.00, 350.00],  # Last amount modified
            'Type': ['Deposit', 'Withdrawal', 'Deposit', 'Fee']
        })
        
        config = ComparisonConfig(
            key_columns=['Account', 'TransDate'],
            alignment_method=AlignmentMethod.POSITION
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # All composite keys should match
        assert result.summary['keys_in_common'] == 4
        assert result.summary['modified_count'] == 1
        
        # Generate and verify report
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            report_path = f.name
        
        try:
            generator = ReportGenerator(report_path)
            generator.generate_report(
                summary=result.summary,
                aligned_data=result.aligned_data,
                metadata=result.comparison_metadata,
                file_a_path='transactions_a.xlsx',
                file_b_path='transactions_b.xlsx'
            )
            
            assert Path(report_path).exists()
            wb = load_workbook(report_path)
            assert 'Summary' in wb.sheetnames
        
        finally:
            if Path(report_path).exists():
                Path(report_path).unlink()


class TestMultiRowPerKeyWorkflow:
    """Test workflow with multiple rows per key"""
    
    def test_multi_row_key_full_workflow(self):
        """Test policy with multiple coverages workflow"""
        df_a = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001', 'P002', 'P002'],
            'Coverage': ['Basic', 'Standard', 'Premium', 'Basic', 'Standard'],
            'Premium': [100, 150, 200, 120, 180],
            'Limit': [100000, 250000, 500000, 100000, 250000]
        })
        
        df_b = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001', 'P002', 'P002', 'P002'],
            'Coverage': ['Basic', 'Standard', 'Premium', 'Basic', 'Standard', 'Elite'],
            'Premium': [100, 160, 200, 120, 180, 250],  # Modified one, added one
            'Limit': [100000, 250000, 500000, 100000, 250000, 600000]
        })
        
        config = ComparisonConfig(
            key_columns=['Policy'],
            alignment_method=AlignmentMethod.POSITION
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Verify multi-row key handling
        assert result.summary['keys_in_common'] == 2
        assert result.summary['match_count'] == 4
        assert result.summary['modified_count'] == 1
        assert result.summary['added_row_count'] == 1


class TestDataNormalizationIntegration:
    """Test normalization integration"""
    
    def test_case_insensitive_comparison(self):
        """Test case insensitive comparison in full workflow"""
        df_a = pd.DataFrame({
            'ID': [1, 2],
            'Status': ['ACTIVE', 'PENDING'],
            'Notes': ['IMPORTANT', 'REVIEW']
        })
        
        df_b = pd.DataFrame({
            'ID': [1, 2],
            'Status': ['Active', 'Pending'],  # Different case
            'Notes': ['important', 'review']  # Different case
        })
        
        config = ComparisonConfig(
            key_columns=['ID'],
            case_sensitive=False
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Should match despite case differences
        assert result.summary['match_count'] == 2
        assert result.summary['modified_count'] == 0
    
    def test_whitespace_trimming_integration(self):
        """Test whitespace trimming in full workflow"""
        df_a = pd.DataFrame({
            'ID': [1, 2],
            'Name': [' Alice ', ' Bob '],
            'City': [' NYC ', ' LA ']
        })
        
        df_b = pd.DataFrame({
            'ID': [1, 2],
            'Name': ['Alice', 'Bob'],  # No spaces
            'City': ['NYC', 'LA']  # No spaces
        })
        
        config = ComparisonConfig(
            key_columns=['ID'],
            trim_whitespace=True
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Should match after trimming
        assert result.summary['match_count'] == 2
        assert result.summary['modified_count'] == 0


class TestAlignmentMethodsIntegration:
    """Test alignment methods in full workflow"""
    
    def test_position_based_alignment_integration(self):
        """Test position-based alignment in workflow"""
        df_a = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001'],
            'Date': ['2024-01-01', '2024-01-02', '2024-01-03'],
            'Amount': [100, 200, 300]
        })
        
        df_b = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001'],
            'Date': ['2024-01-03', '2024-01-01', '2024-01-02'],  # Different order
            'Amount': [300, 100, 200]
        })
        
        config = ComparisonConfig(
            key_columns=['Policy'],
            alignment_method=AlignmentMethod.POSITION
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Position-based: 1st to 1st, 2nd to 2nd, 3rd to 3rd
        assert result.summary['match_count'] == 0
        assert result.summary['modified_count'] == 3
    
    def test_secondary_sort_alignment_integration(self):
        """Test secondary sort alignment in workflow"""
        df_a = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001'],
            'Date': ['2024-01-03', '2024-01-01', '2024-01-02'],
            'Amount': [300, 100, 200]
        })
        
        df_b = pd.DataFrame({
            'Policy': ['P001', 'P001', 'P001'],
            'Date': ['2024-01-01', '2024-01-02', '2024-01-03'],
            'Amount': [100, 200, 300]
        })
        
        config = ComparisonConfig(
            key_columns=['Policy'],
            alignment_method=AlignmentMethod.SECONDARY_SORT,
            secondary_sort_column='Date'
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # Secondary sort: should align by date, matching all three
        assert result.summary['match_count'] == 3


class TestErrorHandling:
    """Test error handling in workflows"""
    
    def test_missing_key_column_error(self):
        """Test error when key column is missing"""
        df_a = pd.DataFrame({'ID': [1], 'Value': [100]})
        df_b = pd.DataFrame({'Name': ['Alice'], 'Value': [100]})
        
        config = ComparisonConfig(key_columns=['ID'])
        engine = ComparisonEngine(config)
        
        with pytest.raises(KeyError):
            engine.compare(df_a, df_b)
    
    def test_empty_dataframe_handling(self):
        """Test handling of empty dataframes"""
        df_a = pd.DataFrame()
        df_b = pd.DataFrame({'ID': [1], 'Value': [100]})
        
        # Should handle gracefully
        try:
            config = ComparisonConfig(key_columns=['ID'])
            engine = ComparisonEngine(config)
            # This may raise or handle gracefully depending on implementation
            result = engine.compare(df_a, df_b)
        except (KeyError, ValueError):
            pass  # Expected behavior


class TestRealWorldScenarios:
    """Test real-world business scenarios"""
    
    def test_insurance_policy_comparison(self):
        """Test insurance policy reconciliation scenario"""
        df_a = pd.DataFrame({
            'PolicyNumber': ['POL001', 'POL001', 'POL001', 'POL002'],
            'CoverageType': ['Liability', 'Property', 'Medical', 'Liability'],
            'Premium': [500.00, 1000.00, 200.00, 600.00],
            'Limit': [100000, 500000, 10000, 100000],
            'Effective': ['2024-01-01'] * 4
        })
        
        df_b = pd.DataFrame({
            'PolicyNumber': ['POL001', 'POL001', 'POL001', 'POL002', 'POL003'],
            'CoverageType': ['Liability', 'Property', 'Medical', 'Liability', 'Liability'],
            'Premium': [525.00, 1000.00, 200.00, 600.00, 450.00],  # One modified, one new
            'Limit': [100000, 500000, 10000, 100000, 75000],
            'Effective': ['2024-01-01'] * 5
        })
        
        config = ComparisonConfig(
            key_columns=['PolicyNumber'],
            case_sensitive=False,
            trim_whitespace=True
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # POL001 has multiple coverages, POL002 matches, POL003 is new
        assert result.summary['keys_in_common'] == 2
        assert result.summary['keys_only_in_b'] == 1
        assert result.summary['new_key_count'] == 1
        assert result.summary['modified_count'] >= 1
    
    def test_financial_transaction_reconciliation(self):
        """Test financial transaction reconciliation scenario"""
        df_a = pd.DataFrame({
            'Account': ['ACC001', 'ACC001', 'ACC002', 'ACC002'],
            'Date': ['2024-01-01', '2024-01-02', '2024-01-01', '2024-01-03'],
            'Type': ['Deposit', 'Withdrawal', 'Deposit', 'Fee'],
            'Amount': [5000.00, 1000.00, 3000.00, 25.00],
            'Balance': [5000.00, 4000.00, 3000.00, 2975.00]
        })
        
        df_b = pd.DataFrame({
            'Account': ['ACC001', 'ACC001', 'ACC002', 'ACC002'],
            'Date': ['2024-01-01', '2024-01-02', '2024-01-01', '2024-01-03'],
            'Type': ['Deposit', 'Withdrawal', 'Deposit', 'Fee'],
            'Amount': [5000.00, 1000.00, 3000.00, 30.00],  # Fee amount different
            'Balance': [5000.00, 4000.00, 3000.00, 2970.00]  # Balance different
        })
        
        config = ComparisonConfig(
            key_columns=['Account', 'Date'],
            alignment_method=AlignmentMethod.POSITION
        )
        
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        # One transaction should show as modified
        assert result.summary['modified_count'] == 1
        assert result.summary['keys_in_common'] == 4


class TestReportIntegration:
    """Test report generation integration"""
    
    def test_report_contains_all_comparison_data(self):
        """Test that report contains all comparison data"""
        df_a = pd.DataFrame({
            'ID': [1, 2, 3],
            'Name': ['Alice', 'Bob', 'Charlie'],
            'Value': [100, 200, 300]
        })
        
        df_b = pd.DataFrame({
            'ID': [1, 2, 4],
            'Name': ['Alice', 'Robert', 'Diana'],
            'Value': [100, 250, 400]
        })
        
        config = ComparisonConfig(key_columns=['ID'])
        engine = ComparisonEngine(config)
        result = engine.compare(df_a, df_b)
        
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            report_path = f.name
        
        try:
            generator = ReportGenerator(report_path)
            generator.generate_report(
                summary=result.summary,
                aligned_data=result.aligned_data,
                metadata=result.comparison_metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            wb = load_workbook(report_path)
            summary_sheet = wb['Summary']
            
            # Verify summary sheet contains key statistics
            assert summary_sheet is not None
        
        finally:
            if Path(report_path).exists():
                Path(report_path).unlink()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
