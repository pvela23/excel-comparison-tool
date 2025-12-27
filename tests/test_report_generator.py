"""
Unit tests for report_generator module
Tests Excel report generation functionality
"""

import pytest
import pandas as pd
import tempfile
from pathlib import Path
from openpyxl import load_workbook
from src.reports.report_generator import ReportGenerator
from src.core import RowStatus


class TestReportGeneratorBasic:
    """Test basic report generation functionality"""
    
    def test_report_generator_creation(self):
        """Test creating a report generator"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            generator = ReportGenerator(f.name)
            assert generator.output_path.name.endswith('.xlsx')
    
    def test_report_generation_empty_data(self):
        """Test generating report with empty data"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 0,
                'total_unique_keys_b': 0,
                'keys_in_common': 0,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 0,
                'match_count': 0,
                'modified_count': 0,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame()
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            assert Path(output_path).exists()
            
            # Verify workbook structure
            wb = load_workbook(output_path)
            assert 'Summary' in wb.sheetnames
            assert 'Legend' in wb.sheetnames
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()
    
    def test_summary_sheet_created(self):
        """Test that summary sheet is created with correct data"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 10,
                'total_unique_keys_b': 12,
                'keys_in_common': 9,
                'keys_only_in_a': 1,
                'keys_only_in_b': 3,
                'total_rows_compared': 25,
                'match_count': 20,
                'modified_count': 3,
                'added_row_count': 1,
                'removed_row_count': 1,
                'new_key_count': 2,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1, 2, 3],
                'A_Value': [100, 200, 300],
                'status': [RowStatus.MATCH.value, RowStatus.MODIFIED.value, RowStatus.ADDED_ROW.value],
                'B_Value': [100, 250, 300]
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='C:\\data\\file_a.xlsx',
                file_b_path='C:\\data\\file_b.xlsx'
            )
            
            # Verify workbook
            wb = load_workbook(output_path)
            assert 'Summary' in wb.sheetnames
            
            summary_sheet = wb['Summary']
            # Check that file paths are in summary
            assert summary_sheet['B4'].value == 'C:\\data\\file_a.xlsx'
            assert summary_sheet['B5'].value == 'C:\\data\\file_b.xlsx'
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()


class TestReportColorCoding:
    """Test color coding in report"""
    
    def test_color_constants_defined(self):
        """Test that color constants are defined"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            generator = ReportGenerator(f.name)
        
        assert 'MATCH' in generator.COLORS
        assert 'MODIFIED' in generator.COLORS
        assert 'ADDED_ROW' in generator.COLORS
        assert 'REMOVED_ROW' in generator.COLORS
        assert 'NEW_KEY' in generator.COLORS
        assert 'REMOVED_KEY' in generator.COLORS
        assert 'HEADER' in generator.COLORS
        
        Path(f.name).unlink()
    
    def test_color_values_valid_hex(self):
        """Test that color values are valid hex codes"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            generator = ReportGenerator(f.name)
        
        for status, color in generator.COLORS.items():
            # Should be 6-character hex code
            assert len(color) == 6
            assert all(c in '0123456789ABCDEF' for c in color.upper())
        
        Path(f.name).unlink()


class TestReportStructure:
    """Test report sheet structure"""
    
    def test_multiple_sheets_created(self):
        """Test that multiple sheets are created"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 5,
                'total_unique_keys_b': 5,
                'keys_in_common': 5,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 5,
                'match_count': 5,
                'modified_count': 0,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1, 2, 3, 4, 5],
                'A_Value': [100, 200, 300, 400, 500],
                'status': [RowStatus.MATCH.value] * 5,
                'B_Value': [100, 200, 300, 400, 500]
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            # Verify sheets exist
            wb = load_workbook(output_path)
            expected_sheets = ['Summary', 'Aligned Diff', 'Legend']
            for sheet in expected_sheets:
                assert sheet in wb.sheetnames
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()
    
    def test_legend_sheet_content(self):
        """Test that legend sheet contains explanation"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 1,
                'total_unique_keys_b': 1,
                'keys_in_common': 1,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 1,
                'match_count': 1,
                'modified_count': 0,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1],
                'A_Value': [100],
                'status': [RowStatus.MATCH.value],
                'B_Value': [100]
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            wb = load_workbook(output_path)
            legend_sheet = wb['Legend']
            assert legend_sheet is not None
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()


class TestReportWithVariousStatuses:
    """Test report generation with various row statuses"""
    
    def test_report_with_all_statuses(self):
        """Test report with all row status types"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 7,
                'total_unique_keys_b': 7,
                'keys_in_common': 5,
                'keys_only_in_a': 2,
                'keys_only_in_b': 2,
                'total_rows_compared': 10,
                'match_count': 3,
                'modified_count': 2,
                'added_row_count': 1,
                'removed_row_count': 1,
                'new_key_count': 2,
                'removed_key_count': 2,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1, 2, 3, 4, 5, 6, 7, 8, 9, 10],
                'A_Value': [100, 200, 300, 400, 500, 600, 700, None, None, None],
                'status': [
                    RowStatus.MATCH.value,
                    RowStatus.MATCH.value,
                    RowStatus.MATCH.value,
                    RowStatus.MODIFIED.value,
                    RowStatus.MODIFIED.value,
                    RowStatus.ADDED_ROW.value,
                    RowStatus.REMOVED_ROW.value,
                    RowStatus.NEW_KEY.value,
                    RowStatus.NEW_KEY.value,
                    RowStatus.REMOVED_KEY.value
                ],
                'B_Value': [100, 200, 300, 450, 550, 800, None, 900, 1000, None]
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            assert Path(output_path).exists()
            wb = load_workbook(output_path)
            assert 'Aligned Diff' in wb.sheetnames
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()
    
    def test_report_with_modified_rows(self):
        """Test report highlighting modified rows"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 3,
                'total_unique_keys_b': 3,
                'keys_in_common': 3,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 3,
                'match_count': 1,
                'modified_count': 2,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1, 2, 3],
                'A_Name': ['Alice', 'Bob', 'Charlie'],
                'A_Value': [100, 200, 300],
                'status': [RowStatus.MATCH.value, RowStatus.MODIFIED.value, RowStatus.MODIFIED.value],
                'B_Name': ['Alice', 'Bobby', 'Charles'],
                'B_Value': [100, 220, 320],
                'changed_cells': ['', 'Name, Value', 'Name, Value']
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            assert Path(output_path).exists()
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()


class TestReportSaving:
    """Test report file saving"""
    
    def test_report_file_saved_successfully(self):
        """Test that report file is saved successfully"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 1,
                'total_unique_keys_b': 1,
                'keys_in_common': 1,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 1,
                'match_count': 1,
                'modified_count': 0,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1],
                'A_Value': [100],
                'status': [RowStatus.MATCH.value],
                'B_Value': [100]
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            # Check file exists and is readable
            assert Path(output_path).exists()
            assert Path(output_path).stat().st_size > 0
            
            # Verify it's a valid Excel file
            wb = load_workbook(output_path)
            assert len(wb.sheetnames) >= 2
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()
    
    def test_report_timestamp_in_path(self):
        """Test that report can use timestamp in path"""
        timestamp = '20240101_120000'
        output_path = f'test_report_{timestamp}.xlsx'
        
        try:
            generator = ReportGenerator(output_path)
            assert timestamp in str(generator.output_path)
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()


class TestReportEdgeCases:
    """Test edge cases in report generation"""
    
    def test_report_with_unicode_characters(self):
        """Test report with unicode characters"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            summary = {
                'total_unique_keys_a': 2,
                'total_unique_keys_b': 2,
                'keys_in_common': 2,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': 2,
                'match_count': 2,
                'modified_count': 0,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': [1, 2],
                'A_Name': ['José', 'François'],
                'status': [RowStatus.MATCH.value, RowStatus.MATCH.value],
                'B_Name': ['José', 'François']
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            assert Path(output_path).exists()
            wb = load_workbook(output_path)
            assert 'Summary' in wb.sheetnames
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()
    
    def test_report_with_large_dataset(self):
        """Test report generation with large dataset"""
        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as f:
            output_path = f.name
        
        try:
            generator = ReportGenerator(output_path)
            
            n = 500
            summary = {
                'total_unique_keys_a': n,
                'total_unique_keys_b': n,
                'keys_in_common': n,
                'keys_only_in_a': 0,
                'keys_only_in_b': 0,
                'total_rows_compared': n,
                'match_count': n - 10,
                'modified_count': 10,
                'added_row_count': 0,
                'removed_row_count': 0,
                'new_key_count': 0,
                'removed_key_count': 0,
            }
            
            aligned_data = pd.DataFrame({
                'key_ID': list(range(n)),
                'A_Value': list(range(100, 100 + n)),
                'status': [RowStatus.MATCH.value] * (n - 10) + [RowStatus.MODIFIED.value] * 10,
                'B_Value': list(range(100, 100 + n))
            })
            
            metadata = {'config': None}
            
            generator.generate_report(
                summary=summary,
                aligned_data=aligned_data,
                metadata=metadata,
                file_a_path='file_a.xlsx',
                file_b_path='file_b.xlsx'
            )
            
            assert Path(output_path).exists()
        
        finally:
            if Path(output_path).exists():
                Path(output_path).unlink()


if __name__ == '__main__':
    pytest.main([__file__, '-v'])
