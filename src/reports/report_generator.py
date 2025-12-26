"""
Excel Report Generator for Comparison Tool
Creates professionally formatted Excel reports with color coding
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any
import pandas as pd
from pathlib import Path


class ReportGenerator:
    """
    Generates formatted Excel comparison reports
    """
   
    # Color definitions
    COLORS = {
        'MATCH': 'FFFFFF',           # White
        'MODIFIED': 'FFF2CC',        # Light Yellow
        'ADDED_ROW': 'C6EFCE',       # Light Green
        'REMOVED_ROW': 'FFC7CE',     # Light Red
        'NEW_KEY': 'DDEBF7',         # Light Blue
        'REMOVED_KEY': 'FCE4D6',     # Light Orange
        'HEADER': '4472C4',          # Blue
        'KEY_SEPARATOR': 'E7E6E6',   # Light Gray
    }
   
    def __init__(self, output_path: str):
        """
        Initialize report generator
       
        Args:
            output_path: Path where the Excel report will be saved
        """
        self.output_path = Path(output_path)
        self.workbook = openpyxl.Workbook()
       
        # Remove default sheet
        if 'Sheet' in self.workbook.sheetnames:
            self.workbook.remove(self.workbook['Sheet'])
   
    def generate_report(
        self,
        summary: Dict[str, Any],
        aligned_data: pd.DataFrame,
        metadata: Dict[str, Any],
        file_a_path: str,
        file_b_path: str
    ):
        """
        Generate complete Excel report
       
        Args:
            summary: Summary statistics dictionary
            aligned_data: Aligned comparison DataFrame
            metadata: Comparison metadata
            file_a_path: Path to File A
            file_b_path: Path to File B
        """
        # Create sheets
        self._create_summary_sheet(summary, file_a_path, file_b_path)
        self._create_aligned_diff_sheet(aligned_data, metadata)
        self._create_legend_sheet(metadata)
       
        # Save workbook
        self.workbook.save(self.output_path)
        print(f"\n✅ Report generated: {self.output_path}")
   
    def _create_summary_sheet(
        self,
        summary: Dict[str, Any],
        file_a_path: str,
        file_b_path: str
    ):
        """Create summary statistics sheet"""
        ws = self.workbook.create_sheet("Summary", 0)
       
        # Title
        ws['A1'] = "Excel Comparison Report - Summary"
        ws['A1'].font = Font(size=16, bold=True, color='FFFFFF')
        ws['A1'].fill = PatternFill(start_color=self.COLORS['HEADER'],
                                     end_color=self.COLORS['HEADER'],
                                     fill_type='solid')
        ws.merge_cells('A1:B1')
       
        # Timestamp
        ws['A2'] = "Generated:"
        ws['B2'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A2'].font = Font(bold=True)
       
        # File information
        row = 4
        ws[f'A{row}'] = "File A:"
        ws[f'B{row}'] = file_a_path
        ws[f'A{row}'].font = Font(bold=True)
       
        row += 1
        ws[f'A{row}'] = "File B:"
        ws[f'B{row}'] = file_b_path
        ws[f'A{row}'].font = Font(bold=True)
       
        # Key statistics header
        row += 2
        ws[f'A{row}'] = "Key Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
       
        # Key statistics
        row += 1
        key_stats = [
            ("Total Unique Keys in File A", summary.get('total_unique_keys_a', 0)),
            ("Total Unique Keys in File B", summary.get('total_unique_keys_b', 0)),
            ("Keys in Common", summary.get('keys_in_common', 0)),
            ("Keys Only in File A", summary.get('keys_only_in_a', 0)),
            ("Keys Only in File B", summary.get('keys_only_in_b', 0)),
        ]
       
        for label, value in key_stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
            row += 1
       
        # Row statistics header
        row += 1
        ws[f'A{row}'] = "Row Statistics"
        ws[f'A{row}'].font = Font(size=14, bold=True)
        ws.merge_cells(f'A{row}:B{row}')
       
        # Row statistics
        row += 1
        row_stats = [
            ("Total Rows Compared", summary.get('total_rows_compared', 0)),
            ("Matching Rows", summary.get('match_count', 0)),
            ("Modified Rows", summary.get('modified_count', 0)),
            ("Added Rows (within shared keys)", summary.get('added_row_count', 0)),
            ("Removed Rows (within shared keys)", summary.get('removed_row_count', 0)),
            ("Rows in New Keys", summary.get('new_key_count', 0)),
            ("Rows in Removed Keys", summary.get('removed_key_count', 0)),
        ]
       
        for label, value in row_stats:
            ws[f'A{row}'] = label
            ws[f'B{row}'] = value
            ws[f'A{row}'].font = Font(bold=True)
           
            # Color code based on status
            if "Modified" in label or "Removed Rows" in label:
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['MODIFIED'],
                                                  end_color=self.COLORS['MODIFIED'],
                                                  fill_type='solid')
            elif "Added" in label:
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['ADDED_ROW'],
                                                  end_color=self.COLORS['ADDED_ROW'],
                                                  fill_type='solid')
            elif "Removed Keys" in label:
                ws[f'B{row}'].fill = PatternFill(start_color=self.COLORS['REMOVED_ROW'],
                                                  end_color=self.COLORS['REMOVED_ROW'],
                                                  fill_type='solid')
           
            row += 1
       
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 35
        ws.column_dimensions['B'].width = 50
   
    def _create_aligned_diff_sheet(
        self,
        aligned_data: pd.DataFrame,
        metadata: Dict[str, Any]
    ):
        """Create aligned diff sheet with color coding"""
        ws = self.workbook.create_sheet("Aligned Diff")
       
        if aligned_data.empty:
            ws['A1'] = "No differences found"
            return
       
        # Prepare data structure
        # Separate key columns, File A columns, status, File B columns
        key_cols = [col for col in aligned_data.columns if col.startswith('key_')]
        a_cols = [col for col in aligned_data.columns if col.startswith('A_')]
        b_cols = [col for col in aligned_data.columns if col.startswith('B_')]
       
        # Create header row
        headers = []
        col_types = []  # Track what type each column is for coloring
       
        # Key columns
        for col in key_cols:
            headers.append(col.replace('key_', '').upper())
            col_types.append('key')
       
        # File A columns
        for col in a_cols:
            headers.append(f"File A: {col.replace('A_', '')}")
            col_types.append('file_a')
       
        # Status column
        headers.append("STATUS")
        col_types.append('status')
       
        # File B columns
        for col in b_cols:
            headers.append(f"File B: {col.replace('B_', '')}")
            col_types.append('file_b')
       
        # Changed cells column (if exists)
        if 'changed_cells' in aligned_data.columns:
            headers.append("CHANGED CELLS")
            col_types.append('changed')
       
        # Write header row
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = PatternFill(start_color=self.COLORS['HEADER'],
                                    end_color=self.COLORS['HEADER'],
                                    fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
       
        # Write data rows
        current_key = None
        border_style = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
       
        for df_row_idx, row in aligned_data.iterrows():
            excel_row = df_row_idx + 2  # +2 because Excel is 1-indexed and we have header
           
            # Check if this is a new key group (for visual separation)
            row_key = tuple(row[col] for col in key_cols)
            is_new_key_group = (current_key != row_key)
            current_key = row_key
           
            col_idx = 1
           
            # Write key columns
            for col in key_cols:
                cell = ws.cell(row=excel_row, column=col_idx, value=row[col])
                cell.border = border_style
                if is_new_key_group:
                    cell.fill = PatternFill(start_color=self.COLORS['KEY_SEPARATOR'],
                                           end_color=self.COLORS['KEY_SEPARATOR'],
                                           fill_type='solid')
                    cell.font = Font(bold=True)
                col_idx += 1
           
            # Write File A columns
            for col in a_cols:
                value = row[col] if col in row and pd.notna(row[col]) else ""
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = border_style
                col_idx += 1
           
            # Write status
            status = row['status']
            cell = ws.cell(row=excel_row, column=col_idx, value=status)
            cell.border = border_style
            cell.alignment = Alignment(horizontal='center')
            cell.font = Font(bold=True)
           
            # Color code based on status
            if status in self.COLORS:
                cell.fill = PatternFill(start_color=self.COLORS[status],
                                       end_color=self.COLORS[status],
                                       fill_type='solid')
            col_idx += 1
           
            # Write File B columns
            for col in b_cols:
                value = row[col] if col in row and pd.notna(row[col]) else ""
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = border_style
               
                # Highlight modified cells
                if status == 'MODIFIED':
                    corresponding_a_col = col.replace('B_', 'A_')
                    if corresponding_a_col in row:
                        a_val = row[corresponding_a_col]
                        b_val = value
                        if pd.notna(a_val) and pd.notna(b_val) and a_val != b_val:
                            cell.fill = PatternFill(start_color=self.COLORS['MODIFIED'],
                                                   end_color=self.COLORS['MODIFIED'],
                                                   fill_type='solid')
               
                col_idx += 1
           
            # Write changed cells info
            if 'changed_cells' in aligned_data.columns:
                value = row['changed_cells'] if pd.notna(row.get('changed_cells')) else ""
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.border = border_style
                cell.font = Font(italic=True, size=9)
       
        # Auto-adjust column widths
        for col_idx in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15
       
        # Freeze header row
        ws.freeze_panes = 'A2'
   
    def _create_legend_sheet(self, metadata: Dict[str, Any]):
        """Create legend/documentation sheet"""
        ws = self.workbook.create_sheet("Legend")
       
        # Title
        ws['A1'] = "Legend & Configuration"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:C1')
       
        # Color legend
        row = 3
        ws[f'A{row}'] = "Color Legend"
        ws[f'A{row}'].font = Font(size=14, bold=True)
       
        row += 1
        ws[f'A{row}'] = "Status"
        ws[f'B{row}'] = "Color"
        ws[f'C{row}'] = "Meaning"
        for col in ['A', 'B', 'C']:
            ws[f'{col}{row}'].font = Font(bold=True)
       
        row += 1
        legend_items = [
            ("MATCH", self.COLORS['MATCH'], "Rows are identical"),
            ("MODIFIED", self.COLORS['MODIFIED'], "Values changed between files"),
            ("ADDED_ROW", self.COLORS['ADDED_ROW'], "Row exists only in File B (within shared key)"),
            ("REMOVED_ROW", self.COLORS['REMOVED_ROW'], "Row exists only in File A (within shared key)"),
            ("NEW_KEY", self.COLORS['NEW_KEY'], "Entire key group only in File B"),
            ("REMOVED_KEY", self.COLORS['REMOVED_KEY'], "Entire key group only in File A"),
        ]
       
        for status, color, meaning in legend_items:
            ws[f'A{row}'] = status
            ws[f'B{row}'] = ""
            ws[f'B{row}'].fill = PatternFill(start_color=color,
                                            end_color=color,
                                            fill_type='solid')
            ws[f'C{row}'] = meaning
            row += 1
       
        # Comparison configuration
        row += 2
        ws[f'A{row}'] = "Comparison Configuration"
        ws[f'A{row}'].font = Font(size=14, bold=True)
       
        row += 1
        config = metadata.get('config')
        if config:
            ws[f'A{row}'] = "Key Columns:"
            ws[f'B{row}'] = ", ".join(config.key_columns)
            ws[f'A{row}'].font = Font(bold=True)
           
            row += 1
            ws[f'A{row}'] = "Alignment Method:"
            ws[f'B{row}'] = config.alignment_method.value
            ws[f'A{row}'].font = Font(bold=True)
           
            if config.secondary_sort_column:
                row += 1
                ws[f'A{row}'] = "Secondary Sort:"
                ws[f'B{row}'] = config.secondary_sort_column
                ws[f'A{row}'].font = Font(bold=True)
           
            row += 1
            ws[f'A{row}'] = "Case Sensitive:"
            ws[f'B{row}'] = "Yes" if config.case_sensitive else "No"
            ws[f'A{row}'].font = Font(bold=True)
           
            row += 1
            ws[f'A{row}'] = "Trim Whitespace:"
            ws[f'B{row}'] = "Yes" if config.trim_whitespace else "No"
            ws[f'A{row}'].font = Font(bold=True)
       
        # Auto-adjust column widths
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 50


# Helper function for quick report generation
def generate_comparison_report(
    output_path: str,
    summary: Dict[str, Any],
    aligned_data: pd.DataFrame,
    metadata: Dict[str, Any],
    file_a_path: str,
    file_b_path: str
):
    """
    Quick helper to generate a comparison report
   
    Args:
        output_path: Where to save the Excel file
        summary: Summary statistics
        aligned_data: Comparison results DataFrame
        metadata: Comparison metadata
        file_a_path: Path to File A
        file_b_path: Path to File B
    """
    generator = ReportGenerator(output_path)
    generator.generate_report(
        summary=summary,
        aligned_data=aligned_data,
        metadata=metadata,
        file_a_path=file_a_path,
        file_b_path=file_b_path
    )